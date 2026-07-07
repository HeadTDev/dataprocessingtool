import csv
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class Processor:
    def process(
        self,
        ksh_path: str,
        matstamm_path: str,
        save_path: str | None = None,
        progress_callback=None,
        is_cancelled=None,
    ):
        try:
            return self._process(
                ksh_path,
                matstamm_path,
                save_path=save_path,
                progress_callback=progress_callback,
                is_cancelled=is_cancelled,
            )
        except InterruptedError:
            return {"cancelled": True, "output_path": None, "row_count": 0}

    def _raise_if_cancelled(self, is_cancelled=None):
        if is_cancelled and is_cancelled():
            raise InterruptedError("A feldolgozás megszakítva.")

    def _process(
        self,
        ksh_path: str,
        matstamm_path: str,
        save_path: str | None = None,
        progress_callback=None,
        is_cancelled=None,
    ):
        project_root = Path(__file__).resolve().parent.parent
        output_dir = project_root / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_csv_path = output_dir / "data.csv"
        output_xlsx_path = output_dir / "data.xlsx"

        if progress_callback:
            progress_callback("KSH fájl beolvasása...", 0, 0)
        with open(ksh_path, "r", encoding="utf-16le", newline="") as f:
            reader = csv.reader(f, delimiter="\t")
            rows = [row for row in reader]

        self._raise_if_cancelled(is_cancelled)
        rows_to_delete = {0, 1, 2, 4, 5}
        cleaned_rows = [row for i, row in enumerate(rows) if i not in rows_to_delete]
        if len(cleaned_rows) < 2:
            raise ValueError(
                "A bemeneti fájl túl kevés sort tartalmaz a feldolgozáshoz."
            )

        header = cleaned_rows[0]
        data_rows = cleaned_rows[1:]

        if progress_callback:
            progress_callback("Matstamm beolvasása...", 0, 0)
        mat_lookup = {}
        wb = openpyxl.load_workbook(matstamm_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("A Matstamm fájlban nincs aktív munkalap.")

        mat_header = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        try:
            anyag_idx = mat_header.index("Anyag")
            beszerzes_idx = mat_header.index("Beszerzés fajtája")
        except ValueError as e:
            raise ValueError(
                "A Matstamm fájl fejlécében nem található 'Anyag' vagy 'Beszerzés fajtája' oszlop."
            ) from e

        for index, row in enumerate(ws.iter_rows(min_row=2), start=1):
            self._raise_if_cancelled(is_cancelled)
            if progress_callback and index % 500 == 0:
                progress_callback("Matstamm sorok feldolgozása...", index, 0)
            anyag = (
                str(row[anyag_idx].value).strip()
                if row[anyag_idx].value is not None
                else ""
            )
            beszerzes = (
                str(row[beszerzes_idx].value).strip()
                if row[beszerzes_idx].value is not None
                else ""
            )
            if anyag:
                mat_lookup[anyag] = beszerzes
        wb.close()

        try:
            data_anyag_idx = header.index("Anyag")
        except ValueError:
            raise ValueError("A KSH fájl fejlécében nincs 'Anyag' oszlop.")

        new_header = header.copy()
        if new_header[-1].strip() != "":
            new_header.append("")
        new_header.append("Iparági értékesítés")

        new_data_rows = []
        total_data_rows = len(data_rows)
        for index, row in enumerate(data_rows, start=1):
            self._raise_if_cancelled(is_cancelled)
            if progress_callback:
                progress_callback("KSH sorok kiegészítése...", index, total_data_rows)
            while len(row) < len(header):
                row.append("")
            anyag_val = str(row[data_anyag_idx]).strip()
            iparagi_ertekesites = mat_lookup.get(anyag_val, "")
            new_row = row.copy()
            if len(new_row) < len(new_header) - 1:
                new_row.append("")
            new_row.append(iparagi_ertekesites)
            new_data_rows.append(new_row)

        def to_clean_float(cell):
            if not cell or str(cell).strip() == "":
                return None
            s = str(cell).replace(" ", "").replace(".", "").replace(",", ".")
            try:
                return float(s)
            except Exception:
                return None

        def find_first_named_index(header_row, name):
            for idx, cell in enumerate(header_row):
                if cell.strip() == name:
                    return idx
            raise ValueError(f"{name} oszlop nem található a fejlécben.")

        jovairas_idx = find_first_named_index(new_header, "Jóváírás")
        penznem_idx = next(
            (
                i
                for i in range(jovairas_idx + 1, len(new_header))
                if new_header[i].strip() == ""
            ),
            None,
        )
        if penznem_idx is None:
            raise ValueError("Nem található üres pénznem oszlop a Jóváírás után.")
        egyenleg_value_idx = penznem_idx + 1

        egysites_header = (
            new_header[:egyenleg_value_idx]
            + ["Egyenleg", ""]
            + new_header[egyenleg_value_idx:]
        )

        forgalom_idx = next(
            (
                i
                for i in range(jovairas_idx - 1, -1, -1)
                if new_header[i].strip() == "Forgalom"
            ),
            None,
        )
        if forgalom_idx is None:
            raise ValueError("Forgalom oszlop nem található a Jóváírás előtt!")
        forgalom_penznem_idx = next(
            (
                i
                for i in range(forgalom_idx + 1, len(new_header))
                if new_header[i].strip() == ""
            ),
            None,
        )
        if forgalom_penznem_idx is None:
            raise ValueError("Forgalom pénznem oszlop nem található!")

        egysites_data_rows = []
        total_rows = len(new_data_rows)
        for index, row in enumerate(new_data_rows, start=1):
            self._raise_if_cancelled(is_cancelled)
            if progress_callback:
                progress_callback("Egyenleg számítása...", index, total_rows)
            while len(row) < len(new_header):
                row.append("")
            jovairas_value = row[jovairas_idx]
            jovairas_currency = row[penznem_idx]
            forgalom_value = row[forgalom_idx]
            forgalom_currency = row[forgalom_penznem_idx]
            forgalom_val = to_clean_float(forgalom_value)
            jovairas_val = to_clean_float(jovairas_value)
            egyenleg_val = (forgalom_val or 0.0) + (jovairas_val or 0.0)
            egyenleg_cur = forgalom_currency or jovairas_currency or ""
            new_row = (
                row[:egyenleg_value_idx]
                + [f"{egyenleg_val:.2f}", egyenleg_cur]
                + row[egyenleg_value_idx:-2]
                + [row[-2], row[-1]]
            )
            egysites_data_rows.append(new_row)

        if progress_callback:
            progress_callback("CSV mentése...", 0, 0)
        with open(output_csv_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(egysites_header)
            writer.writerows(egysites_data_rows)

        if progress_callback:
            progress_callback("XLSX létrehozása...", 0, 0)
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("Nem sikerült XLSX munkalapot létrehozni.")
        ws.title = "Adatok"

        egyenleg_col_idx = None
        iparagi_col_idx = None
        for idx, name in enumerate(egysites_header):
            if name.strip() == "Egyenleg":
                egyenleg_col_idx = idx
            if name.strip() == "Iparági értékesítés":
                iparagi_col_idx = idx

        ws.append(egysites_header)
        for index, row in enumerate(egysites_data_rows, start=1):
            self._raise_if_cancelled(is_cancelled)
            if progress_callback:
                progress_callback("Excel sorok írása...", index, total_rows)
            excel_row = []
            for idx, cell in enumerate(row):
                if idx == egyenleg_col_idx:
                    try:
                        excel_row.append(round(float(cell), 2))
                    except Exception:
                        excel_row.append(cell)
                elif idx == iparagi_col_idx:
                    excel_row.append(cell)
                else:
                    num = to_clean_float(cell)
                    excel_row.append(num if num is not None else cell)
            ws.append(excel_row)

        max_row = ws.max_row
        ws.auto_filter.ref = ws.dimensions

        highlight_names = ["Forgalom", "Jóváírás", "Egyenleg", "Iparági értékesítés"]
        highlight_cols = [
            idx + 1
            for idx, name in enumerate(egysites_header)
            if name.strip() in highlight_names
        ]

        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        for col in highlight_cols:
            self._raise_if_cancelled(is_cancelled)
            for row_num in range(2, max_row + 1):
                ws.cell(row=row_num, column=col).fill = yellow_fill

        if progress_callback:
            progress_callback("Oszlopszélességek beállítása...", 0, 0)
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(int(col[0].column or 1))
            for cell in col:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            ws.column_dimensions[column].width = max_length + 2

        final_output_path = Path(save_path) if save_path else output_xlsx_path
        if progress_callback:
            progress_callback("XLSX mentése...", 0, 0)
        wb.save(final_output_path)

        cleanup_message = ""
        if save_path:
            try:
                shutil.rmtree(output_dir)
                cleanup_message = "Az output mappa törlésre került."
            except Exception as exc:
                cleanup_message = f"Figyelem, az output mappa törlése sikertelen: {exc}"

        return {
            "cancelled": False,
            "output_path": str(final_output_path),
            "row_count": total_rows,
            "cleanup_message": cleanup_message,
        }
