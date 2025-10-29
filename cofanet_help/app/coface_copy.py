"""Coface Excel filling module with fuzzy company name matching."""

import os
import csv
import sys
import logging
from typing import List, Dict, Tuple, Optional
from functools import lru_cache
from pathlib import Path

from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import numbers

from common import Config

logger = logging.getLogger("cofanet_help")

# Pre-compile regex for company type removal
import unicodedata
import re

COMPANY_TYPES_PATTERN = re.compile(
    r'\b(kft\.?|zrt\.?|bt\.?|gmbh|ltd\.?|nyrt\.?|rt\.?|b\.v\.?|s\.a\.?|s\.r\.l\.?|se|ev|sas|'
    r'korlatolt felelossegu tarsasag|szolgaltato|kereskedelmi|beteti tarsasag|nonprofit)\b'
)
NON_ALPHANUMERIC_PATTERN = re.compile(r'[^a-z0-9]')
WHITESPACE_PATTERN = re.compile(r'\s+')


@lru_cache(maxsize=512)
def normalize_name(name: str) -> str:
    """
    Normalize company name for matching (cached for performance).
    
    Args:
        name: Company name to normalize
        
    Returns:
        Normalized company name
    """
    if not name:
        return ""
    
    # Remove accents
    name = unicodedata.normalize('NFKD', name).encode('ASCII', 'ignore').decode()
    name = name.lower()
    
    # Remove company types
    name = COMPANY_TYPES_PATTERN.sub('', name)
    
    # Remove non-alphanumeric characters
    name = NON_ALPHANUMERIC_PATTERN.sub(' ', name)
    
    # Normalize whitespace
    name = WHITESPACE_PATTERN.sub(' ', name).strip()
    
    return name

def best_fuzzy_match(
    vevo_name: str,
    coface_names: List[str],
    threshold: float = None
) -> Optional[int]:
    """
    Find best fuzzy match for company name using rapidfuzz (much faster than SequenceMatcher).
    
    Args:
        vevo_name: Customer name to match
        coface_names: List of Coface company names
        threshold: Matching threshold (0-100), defaults to Config value
        
    Returns:
        Index of best match or None if no match above threshold
    """
    if threshold is None:
        threshold = Config.FUZZY_MATCH_THRESHOLD * 100  # rapidfuzz uses 0-100 scale
    
    norm_vevo = normalize_name(vevo_name)
    
    # Pre-normalize all coface names (can be done once and cached)
    norm_coface_names = [normalize_name(name) for name in coface_names]
    
    # Use rapidfuzz's extractOne for best match (C-based, very fast)
    result = process.extractOne(
        norm_vevo,
        norm_coface_names,
        scorer=fuzz.WRatio  # Weighted ratio - best for company names
    )
    
    if result and result[1] >= threshold:
        return result[2]  # Return index
    
    return None

def format_amount(amount_str):
    """
    Formázza az összeget helyes számformátumra:
    - ezres elválasztó: vessző
    - tizedes jel: pont
    - excel cellába számként kerül be
    """
    if not amount_str:
        return None, ""
    amt = amount_str.replace(" ", "")
    if ',' in amt and amt.count(',') == 1:
        amt = amt.replace('.', '')
        amt = amt.replace(',', '.')
    elif amt.count('.') > 1 and ',' not in amt:
        parts = amt.split('.')
        tizedes = parts[-1]
        amt = ''.join(parts[:-1]) + '.' + tizedes
    try:
        num = float(amt)
        # Sztring formázás: 1,234,567.89
        str_formatted = f"{num:,.2f}"
        return num, str_formatted
    except Exception:
        return None, amount_str

def find_row_for_company(vevo_name: str, coface_names: List[str]) -> Optional[int]:
    """
    Find row index for a company name in Coface names list.
    
    Tries exact match first, then fuzzy match, then first word match.
    
    Args:
        vevo_name: Customer name to find
        coface_names: List of Coface company names
        
    Returns:
        Index of matching row or None
    """
    vevo_name_lower = vevo_name.lower().strip()
    
    # Try exact match first
    for idx, name in enumerate(coface_names):
        if vevo_name_lower == name.lower().strip():
            logger.debug(f"Exact match found for '{vevo_name}' at index {idx}")
            return idx
    
    # Try fuzzy match
    idx = best_fuzzy_match(vevo_name, coface_names, threshold=80)
    if idx is not None:
        logger.debug(f"Fuzzy match found for '{vevo_name}' at index {idx}: {coface_names[idx]}")
        return idx
    
    # Try first word match
    vevo_first_word = vevo_name_lower.split()[0] if vevo_name_lower.split() else ""
    if vevo_first_word:
        for idx, name in enumerate(coface_names):
            coface_words = name.lower().split()
            if coface_words and vevo_first_word == coface_words[0]:
                logger.debug(f"First word match found for '{vevo_name}' at index {idx}")
                return idx
    
    logger.warning(f"No match found for customer: '{vevo_name}'")
    return None

def fill_coface_excel_and_open(
    coface_excel_path: str,
    vevok_csv_path: str,
    save_path: Optional[str] = None
) -> str:
    """
    Fill Coface Excel with customer data and open it.
    
    Args:
        coface_excel_path: Path to Coface Excel template
        vevok_csv_path: Path to customer data CSV
        save_path: Output path (defaults to same directory as template)
        
    Returns:
        Path to saved Excel file
    """
    logger.info(f"Filling Coface Excel: {coface_excel_path}")
    
    # Read customer data
    with open(vevok_csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        vevok_data = [(row.get("Vevő", "").strip(), row.get("Forintosítva HUF", "")) for row in reader]
    
    logger.info(f"Loaded {len(vevok_data)} customer records")

    wb = load_workbook(coface_excel_path)
    ws = wb.active
    if ws is None:
        raise Exception("Nem sikerült megnyitni az aktív munkalapot!")

    header_row_idx, header = None, []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10), 1):
        header = [str(cell.value).strip() if cell.value else "" for cell in row]
        if "Cégnév" in header and "Számlázott összeg" in header:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise Exception("Nem található 'Cégnév' és 'Számlázott összeg' fejléc!")

    cegnev_col = header.index("Cégnév")
    osszeg_col = header.index("Számlázott összeg")

    coface_rows = list(ws.iter_rows(min_row=header_row_idx+1, max_row=ws.max_row))
    coface_names = [
        str(row[cegnev_col].value or "").strip()
        for row in coface_rows
    ]

    from openpyxl.cell.cell import MergedCell
    for vevo_name, amount in vevok_data:
        if not vevo_name:
            continue
        idx = find_row_for_company(vevo_name, coface_names)
        if idx is not None:
            target_row = coface_rows[idx]
            cell = target_row[osszeg_col]
            if not isinstance(cell, MergedCell):
                num_value, str_formatted = format_amount(amount)
                if num_value is not None:
                    cell.value = num_value
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1  # '1,234,567.89'
                else:
                    cell.value = amount

    # --- MENTÉS FELHASZNÁLÓ ÁLTAL VÁLASZTOTT HELYRE ---
    if save_path is None:
        output_path = os.path.join(os.path.dirname(coface_excel_path), "coface_output.xlsx")
    else:
        output_path = save_path
    wb.save(output_path)
    try:
        if sys.platform.startswith('darwin'):
            os.system(f'open "{output_path}"')
        elif os.name == 'nt':
            os.startfile(output_path)
        else:
            os.system(f'xdg-open "{output_path}"')
    except Exception:
        pass

    return output_path